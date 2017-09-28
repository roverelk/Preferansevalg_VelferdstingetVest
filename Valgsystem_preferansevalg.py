    # Valgsystsem for VT
# Preferansevalg fr n antall deltakere og for m antall stemmer.

# Det kommer inn stemmer i form av et excel-ark.
# Stemmene telles opp.

# Prosessen skrives til .txt-fil som viser alle avgjøresler som tas

# Endelig rangert resultat skrives til en egen .txt-fil.

# LES INN STEMMER
from openpyxl import Workbook, load_workbook
from Preferansevalg_imports_v3 import importer_stemmesedler, finn_sperregrense, noen_har_vunnet, forste_opptelling, \
    kontroll_antall_kandidater_igjen, ny_fordeling_etter_vunnet, ny_fordeling_etter_tap, print_status

# EKSPORTER RESULTATER
output_status = 'Status.txt'
output_file = 'Resultat.xlsx'
# Sett opp excel-ark for å skrive inn data

wb = Workbook()
ws = wb.active
R = 1
C = 1
ws.cell(row=R, column=C).value = 'Plass'
ws.cell(row=R, column=C+1).value = 'Kandidater'
R += 1
wb.save(output_file)


# IMPORTER STEMMESEDLER
input_file = 'Stemmesedler.xlsx'
kandidater, stemmesedler = importer_stemmesedler(input_file)

# SETT SPERREGRENSE
tekst = 'Antall kandidater: ' + str(len(kandidater))
print_status(output_status, tekst)
tekst = '\nAntall stemmer: ' + str(len(stemmesedler))
print_status(output_status, tekst)
sperregrense = finn_sperregrense(kandidater, stemmesedler)
tekst = '\nSperregrense: ' + str(sperregrense)
print_status(output_status, tekst)

score = forste_opptelling(stemmesedler, kandidater, output_status)

tapende_plass = len(kandidater)

# SE OM NOEN HAR VUNNET
flere_kandidater = True
while(flere_kandidater):
    tekst = '\n\nRunde nummer ' + str(R - 1)
    print_status(output_status, tekst)

    if noen_har_vunnet(stemmesedler, sperregrense, score):
        tekst = '\nKandidat har VUNNET'
        print_status(output_status, tekst)

        score, stemmesedler, choose_cand = ny_fordeling_etter_vunnet(score, stemmesedler, sperregrense, kandidater,
                                                                     output_status)

        wb = load_workbook(output_file)
        ws = wb.active
        ws.cell(row=R, column=C).value = R - 1
        ws.cell(row=R, column=C+1).value = str(kandidater[choose_cand])
        R += 1
        wb.save(output_file)
    else:
        # FINN PERSON MED LAVEST SCORE
        tekst = '\nKandidat har TAPT'
        print_status(output_status, tekst)

        score, stemmesedler, choose_cand = ny_fordeling_etter_tap(score, stemmesedler, sperregrense, kandidater,
                                                                  output_status)
        wb = load_workbook(output_file)
        ws = wb.active
        ws.cell(row=R, column=C).value = tapende_plass
        ws.cell(row=R, column=C + 1).value = str(kandidater[choose_cand])
        R += 1
        wb.save(output_file)
        tapende_plass -= 1
    # Tell opp hvor mange stemmer hver person fikk som sin førstestemme.
    flere_kandidater = kontroll_antall_kandidater_igjen(score)
    print('Antall kandidater igjen:', flere_kandidater)

    # Skrive sistemann på arket

tekst = '\n\nÅpne Resultat.xlsx for å se endelig resultat.'
print_status(output_status, tekst)



