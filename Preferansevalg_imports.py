# -*- coding: utf-8 -*-
# Alle delprogrammer som skal brukes av hovedprogrammet

from openpyxl import load_workbook
from random import randint

from Preferansevalg_2vinnere import flere_vinnere

def print_status(file, tekst):
    with open(file, "a") as my_file:
        my_file.write(tekst)


def importer_stemmesedler(input_file):
    wb = load_workbook(input_file)
    ws = wb.active

    # Get names of all candidates
    candidates = []
    more_candidates = True
    R = 2
    C = 1

    while more_candidates:
        cand = ws.cell(row=R, column=C).value
        if cand is not None:
            candidates.append(cand)
            R = R + 1
        else:
            more_candidates = False

    # Get all votes
    ballots = []
    one_ballot = []
    more_ballots = True
    R = 2
    C = 2

    while more_ballots:
        bal = ws.cell(row=R, column=C).value
        if bal is not None:
            one_ballot.append(bal)
            R = R + 1
        else:
            R = 2
            C = C + 1
            ballots.append(one_ballot)
            one_ballot = []

            if ws.cell(row=R, column=C).value is None:
                more_ballots = False

    #Sett alle blanke stemmer til -1
    for i in range(len(ballots)):
        for j in range(len(ballots[i])):
            if ballots[i][j] == 0:
                ballots[i][j] = -1

    kandidater = candidates
    stemmesedler = ballots

    return_tuplet = kandidater, stemmesedler
    return return_tuplet


def finn_sperregrense(antall_kandidater, antall_stemmesedler):
    return (len(antall_stemmesedler) / len(antall_kandidater)) + 0.01


def forste_opptelling(stemmesedler, kandidater, out_status):
    first_count = []
    score = []
    for i in range(len(kandidater)):
        score.append(0)

    for i in range(len(stemmesedler)):
        for j in range(len(stemmesedler[0])):
            if stemmesedler[i][j] is 1:
                score[j] = score[j] + 1
                stemmesedler[i][j] = -2

    tekst = '\n\nFørste opptelling.\nKandidater med score:'
    print_status(out_status, tekst)

    for i in range(len(kandidater)):
        name_spacing = ' ' * (len(max(kandidater, key=len)) - len(kandidater[i]))
        tekst = str('\n' + kandidater[i]) + name_spacing + '\t:\t' + str(score[i])
        print_status(out_status, tekst)

    return score


def noen_har_vunnet(stemmesedler, sperregrense, score):
    for i in range(len(stemmesedler[0])):
        if score[i] > sperregrense:
            return True
    return False


def kontroll_antall_kandidater_igjen(score):
    antall_kandidater_igjen = 0
    for i in range(len(score)):
        if score[i] >= 0:
            antall_kandidater_igjen += 1
    if antall_kandidater_igjen >= 1:
        return True
    else:
        return False


def ny_fordeling_etter_vunnet(score, stemmesedler, sperregrense, kandidater, out_status):

    def hoyest_antall_stemmer(kandidater_, score_, high_score_):
        for i in range(len(kandidater_)):
            if score_[i] > high_score_:
                high_score_ = score_[i]
        return high_score_

    def kandidat_med_flest_stemmer(kandidater_, high_score_, score_, stemmesedler_):
        cand_over = []

        # Velg flere kandidater om de alle har like mange og flest poeng
        for i_ in range(len(kandidater_)):
            if score[i_] == high_score_:
                cand_over.append(i_)

        # Hvis flere enn 1 velg en random
        if len(cand_over) > 1:
            choose_cand_ = flere_vinnere(cand_over, score_, stemmesedler_, kandidater_)
            print(choose_cand_)
        else:
            choose_cand_ = cand_over[0]

        tekst = '\n' + str(kandidater_[choose_cand_]) + ' har vunnet med ' + str(round(high_score_, 2)) + ' stemmer.'
        print_status(out_status, tekst)
        return choose_cand_

    def ny_fordeling(score_, choose_cand_, sperregrense_, kandidater_, stemmesedler_):

        antall_stemmer_refordeles = 0

        # Ta første persom som har vunnet og fordel dens stemmer på nytt.
        fordeling = []
        fordel_fordeling = []
        for i in range(len(kandidater_)):
            fordeling.append(0)
            fordel_fordeling.append(0)

        # Se gjennom alle stemmesedlene
        for i in range(len(stemmesedler_)):
            # Hvis stemmeseddel i, som har 1 i samme posisjon som vinnende kandidat
            if stemmesedler_[i][choose_cand_] == -2:
                # Tell antall stemmer som skal refordeles
                antall_stemmer_refordeles += 1

                # Se gjennom alle posisjoner i den stemmeseddelen
                # Nullstill den vinnende stemmens posisjon
                stemmesedler_[i][choose_cand_] = -1

                leter_etter_neste_preferanse = True
                preferanse = 1
                posisjon = 0

                while(leter_etter_neste_preferanse):
                    if stemmesedler_[i][posisjon] == preferanse and score_[posisjon] != -1:
                        fordeling[posisjon] += 1
                        leter_etter_neste_preferanse = False

                        # Sett den stemmeseddelen til å bli satt som brukt for sin posisjon, så den ikke kan gå til
                        # samme person en gang til.
                        stemmesedler_[i][posisjon] = -1
                    else:
                        posisjon += 1

                        if posisjon >= len(stemmesedler_[i]):
                            posisjon = 0
                            preferanse += 1

                            if preferanse >= len(stemmesedler_[i]):
                                posisjon = 0
                                preferanse = 1
                                leter_etter_neste_preferanse = False

        score_gammel = []
        for i in range(len(score_)):
            score_gammel.append(score_[i])

        if antall_stemmer_refordeles == 0:
            print('Ingen stemmer å fordele :-(')
        else:
            fordelingsnokkel = (score_[choose_cand_] - sperregrense_) / antall_stemmer_refordeles

            for i in range(len(fordeling)):
                fordel_fordeling[i] = fordeling[i] * fordelingsnokkel

            for i in range(len(score_)):
                score_[i] = score_[i] + fordel_fordeling[i]

        score[choose_cand_] = -1.0

        tekst = "\nStemmene til " + str(kandidater[choose_cand_] + " er fordelt:")
        print_status(out_status, tekst)

        for i in range(len(score_)):
            if score_[i] > 0:
                name_spacing = ' ' * (len(max(kandidater, key=len))-len(kandidater[i]))

                tekst = '\n' + str(kandidater[i]) + name_spacing + '\t:\t' + str(round(score_gammel[i], 2)) + '\t+\t' \
                    + str(round(fordel_fordeling[i], 2)) + '\t=\t' + str(round(score_[i], 2))
                print_status(out_status, tekst)

        return score_

    # Finn høyest antall stemmer
    high_score = 0
    high_score = hoyest_antall_stemmer(kandidater, score, high_score)

    # Finn hvem som har høyest antal stemmer
    choose_cand = kandidat_med_flest_stemmer(kandidater, high_score, score, stemmesedler)

    # Fordel overflødige stemmer fra kandidaten som har vunnet
    score = ny_fordeling(score, choose_cand, sperregrense, kandidater, stemmesedler)

    return score, stemmesedler, choose_cand


def ny_fordeling_etter_tap(score, stemmesedler, sperregrense, kandidater, out_status):

    def laveste_antall_stemmer(kandidater_, score_, low_score_):
        for i in range(len(kandidater_)):
            if score_[i] < low_score_ and score_[i] >= 0:
                low_score_ = score_[i]
        return low_score_

    def kandidat_med_ferrest_stemmer(kandidater_, low_score_):
        cand_under = []

        for i_ in range(len(kandidater_)):
            if score[i_] == low_score_:
                cand_under.append(i_)

        # Hvis flere enn 1 velg en random
        if len(cand_under) > 1:
            choose_cand_ = cand_under[randint(0, len(cand_under) - 1)]
            print(choose_cand_)
        else:
            choose_cand_ = cand_under[0]

        tekst = '\n' + str(kandidater_[choose_cand_]) + ' har tapt med ' + str(round(low_score_, 2)) + ' stemmer.'
        print_status(out_status, tekst)
        return choose_cand_

    def ny_fordeling(score_, choose_cand_, sperregrense_, kandidater_, stemmesedler_):

        antall_stemmer_refordeles = 0

        # Ta første persom som har vunnet og fordel dens stemmer på nytt.
        fordeling = []
        fordel_fordeling = []
        for i in range(len(kandidater_)):
            fordeling.append(0)
            fordel_fordeling.append(0)

        # Se gjennom alle stemmesedlene
        for i in range(len(stemmesedler_)):


            # Hvis stemmeseddel i, som har 1 i samme posisjon som vinnende kandidat
            if stemmesedler_[i][choose_cand_] == -2:
                # Tell antall stemmer som skal refordeles
                antall_stemmer_refordeles += 1

                # Se gjennom alle posisjoner i den stemmeseddelen
                # Nullstill den vinnende stemmens posisjon
                stemmesedler_[i][choose_cand_] = -1

                leter_etter_neste_preferanse = True
                preferanse = 1
                posisjon = 0

                while(leter_etter_neste_preferanse):
                    if stemmesedler_[i][posisjon] == preferanse and score_[posisjon] != -1:
                        fordeling[posisjon] += 1
                        leter_etter_neste_preferanse = False

                    else:
                        posisjon += 1

                        if posisjon >= len(stemmesedler_[i]):
                            posisjon = 0
                            preferanse += 1

                            if preferanse >= len(stemmesedler_[i]):
                                posisjon = 0
                                preferanse = 1
                                leter_etter_neste_preferanse = False

        for i in range(len(score_)):
            score_[i] = score_[i] + fordeling[i]
        score[choose_cand_] = -1.0
        return score_

    # Finn lavest antall stemmer
    low_score = 10
    low_score = laveste_antall_stemmer(kandidater, score, low_score)

    # Finn hvem som har lavest antall stemmer
    choose_cand = kandidat_med_ferrest_stemmer(kandidater, low_score)

    # Fordel overflødige stemmer fra kandidaten til de resterende kandidatene
    score = ny_fordeling(score, choose_cand, sperregrense, kandidater, stemmesedler)

    return score, stemmesedler, choose_cand