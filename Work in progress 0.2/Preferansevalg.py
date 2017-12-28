# -*- coding: utf-8 -*-
# Valgsystsem for VT
# Preferansevalg fr n antall deltakere og for m antall stemmer.

# Det kommer inn stemmer i form av et excel-ark.
# Stemmene telles opp.

# Prosessen skrives til .txt-fil som viser alle avgjøresler som tas

# Endelig rangert resultat skrives til en egen .txt-fil.

# LES INN STEMMER
from Preferansevalg_vedlegg import importer_stemmesedler, \
    finn_sperregrense, \
    noen_har_vunnet, \
    forste_opptelling, \
    kontroll_antall_kandidater_igjen, \
    ny_fordeling_etter_vunnet, \
    ny_fordeling_etter_tap, \
    print_status, \
    alle_andre_opptellinger

from openpyxl import Workbook, load_workbook
import time
import os
import shutil
import numpy as np

# SET TIME
t = time.localtime()

# SET OUTPUT FOLDER
output_file =        'Resultat.xlsx'
output_status =      'Status.txt'
output_deep_status = 'Deep_status.txt'

# output_result_array = []

# Sett opp excel-ark for å skrive inn data
wb = Workbook()
ws = wb.active
R = 1
C = 1
ws.cell(row=R, column=C).value = 'Plass'
ws.cell(row=R, column=C+1).value = 'Kandidater'
R += 1
wb.save(output_file)


# IMPORTER STEMMESEDLER
input_file = 'Stemmesedler.xlsx'
kandidater, stemmesedler = importer_stemmesedler(input_file)

# SETT SPERREGRENSE
sperregrense = finn_sperregrense(kandidater, stemmesedler)

print('Sperregrense: ' + str(sperregrense))

# PRINT STATUS TIL TXT
print_status(output_status, ('Antall kandidater: ' + str(len(kandidater)) +
                             '\nAntall stemmer: ' + str(len(stemmesedler)) +
                             '\nSperregrense: ' + str(sperregrense)))


score, oversikt = forste_opptelling(stemmesedler, kandidater, output_status, output_deep_status)

resulterende_rangering = np.zeros(shape=np.shape(kandidater))
tapende_plass = len(kandidater)
output_result_array = ["" for x in range(len(kandidater))]

print('--- SCORE ---')
print(score)

# SE OM NOEN HAR VUNNET
flere_kandidater = True
while(flere_kandidater):
    print('Det er flere kandidater')
    if noen_har_vunnet(stemmesedler, sperregrense, score):
        # Noen vinner
        score, oversikt, resulterende_rangering = ny_fordeling_etter_vunnet(score, stemmesedler, sperregrense,
                                                                            resulterende_rangering,
                                                                            kandidater, oversikt)
    else:
        # Noen taper
        score, oversikt, resulterende_rangering = ny_fordeling_etter_tap(score, stemmesedler, sperregrense,
                                                                         resulterende_rangering,
                                                                         kandidater, oversikt)
    print('--- RESULTAT ---')
    print(resulterende_rangering)
    print('--- SCORE ---')
    print(score)
    oversikt = alle_andre_opptellinger(stemmesedler, oversikt, score, sperregrense)

#tekst = '\n\nÅpne Resultat.xlsx for å se endelig resultat.'
#print_status(output_status, tekst)
#print_status(output_deep_status, tekst)


# EKSPORTER RESULTATER + SORT CORRECTLY
#output_folder = 'Results [' + str(time.strftime('%y-%m-%d] [%H-%M-%S]'))
#try:
#    os.stat(output_folder)
#except:
#    os.mkdir(output_folder)
##Move all the files into folder
#os.renames(output_file,         output_folder + '/1 - ' + output_file)
#os.renames(output_status,       output_folder + '/2 - ' + output_status)
#os.renames(output_deep_status,  output_folder + '/3 - ' + output_deep_status)
#shutil.copyfile(input_file,     output_folder + '/4 - ' + input_file)

#print(output_result_array)

