# -*- coding: utf-8 -*-
# Valgsystsem for VT
# Preferansevalg fr n antall deltakere og for m antall stemmer.

# Det kommer inn stemmer i form av et excel-ark.
# Stemmene telles opp.

# Prosessen skrives til .txt-fil som viser alle avgjøresler som tas

# Endelig rangert resultat skrives til en egen .txt-fil.

# LES INN STEMMER
from Preferansevalg_vedlegg import importer_stemmesedler, finn_sperregrense, noen_har_vunnet, forste_opptelling, \
    kontroll_antall_kandidater_igjen, ny_fordeling_etter_vunnet, ny_fordeling_etter_tap, print_status

from openpyxl import Workbook, load_workbook
import time
import os
import shutil
import numpy as np

# SET TIME
t = time.localtime()

# SET OUTPUT FOLDER
output_file =        'Resultat.xlsx'
output_status =      'Status.txt'
output_deep_status = 'Deep_status.txt'

# output_result_array = []

# Sett opp excel-ark for å skrive inn data
wb = Workbook()
ws = wb.active
R = 1
C = 1
ws.cell(row=R, column=C).value = 'Plass'
ws.cell(row=R, column=C+1).value = 'Kandidater'
R += 1
wb.save(output_file)


# IMPORTER STEMMESEDLER
input_file = 'Stemmesedler.xlsx'
kandidater, stemmesedler = importer_stemmesedler(input_file)

# SETT SPERREGRENSE
sperregrense = finn_sperregrense(kandidater, stemmesedler)

# PRINT STATUS TIL TXT
print_status(output_status, ('Antall kandidater: ' + str(len(kandidater)) +
                             '\nAntall stemmer: ' + str(len(stemmesedler)) +
                             '\nSperregrense: ' + str(sperregrense)))

print_status(output_deep_status, ('Antall kandidater: ' + str(len(kandidater)) +
                                  '\nAntall stemmer: ' + str(len(stemmesedler)) +
                                  '\nSperregrense: ' + str(sperregrense)))

score = forste_opptelling(stemmesedler, kandidater, output_status, output_deep_status)

tapende_plass = len(kandidater)
output_result_array = ["" for x in range(len(kandidater))]

# SE OM NOEN HAR VUNNET
flere_kandidater = True
while(flere_kandidater):
    print_status(output_status, ('\n\nRunde nummer ' + str(R - 1)))
    print_status(output_deep_status, ('\n\nRunde nummer ' + str(R - 1)))

    if noen_har_vunnet(stemmesedler, sperregrense, score):
        print_status(output_status, '\nKandidat har VUNNET')
        print_status(output_deep_status, '\nKandidat har VUNNET')

        tekst = '\nFØR FORDELING\nSCORE\n' + str(score) + '\nSTEMMESEDLER\n' + str(stemmesedler)
        print_status(output_deep_status, tekst)

        score, stemmesedler, choose_cand = ny_fordeling_etter_vunnet(score, stemmesedler, sperregrense, kandidater,
                                                                     output_status, output_deep_status)

        tekst = '\nFØR FORDELING\nSCORE\n' + str(score) + '\nSTEMMESEDLER\n' + str(stemmesedler)
        print_status(output_deep_status, tekst)

        wb = load_workbook(output_file)
        ws = wb.active
        ws.cell(row=R, column=C).value = R - 1
        ws.cell(row=R, column=C+1).value = str(kandidater[choose_cand])
        wb.save(output_file)

        print('R:' + str(R))
        output_result_array[R - 2] = str(kandidater[choose_cand])

        R += 1

    else:
        # FINN PERSON MED LAVEST SCORE
        print_status(output_status, '\nKandidat har TAPT')
        print_status(output_deep_status, '\nKandidat har TAPT')

        tekst = '\nFØR FORDELING\nSCORE\n' + str(score) + '\nSTEMMESEDLER\n' + str(stemmesedler)
        print_status(output_deep_status, tekst)

        score, stemmesedler, choose_cand = ny_fordeling_etter_tap(score, stemmesedler, sperregrense, kandidater,
                                                                  output_status, output_deep_status)

        tekst = '\nFØR FORDELING\nSCORE\n' + str(score) + '\nSTEMMESEDLER\n' + str(stemmesedler)
        print_status(output_deep_status, tekst)

        wb = load_workbook(output_file)
        ws = wb.active
        ws.cell(row=R, column=C).value = tapende_plass
        ws.cell(row=R, column=C + 1).value = str(kandidater[choose_cand])
        R += 1
        wb.save(output_file)

        output_result_array[tapende_plass - 1] = str(kandidater[choose_cand])

        tapende_plass -= 1

    # Tell opp hvor mange stemmer hver person fikk som sin førstestemme.
    flere_kandidater = kontroll_antall_kandidater_igjen(score)
    print('Antall kandidater igjen:', flere_kandidater)

    # Skrive sistemann på arket
    print_status(output_deep_status, '\nOversikt score mellom runder')
    for i in range(len(kandidater)):
        name_spacing = ' ' * (len(max(kandidater, key=len)) - len(kandidater[i]))
        tekst = str('\n' + kandidater[i]) + name_spacing + '\t:\t' + str(score[i])
        print_status(output_deep_status, tekst)

tekst = '\n\nÅpne Resultat.xlsx for å se endelig resultat.'
print_status(output_status, tekst)
print_status(output_deep_status, tekst)


# EKSPORTER RESULTATER + SORT CORRECTLY
output_folder = 'Results [' + str(time.strftime('%y-%m-%d] [%H-%M-%S]'))
try:
    os.stat(output_folder)
except:
    os.mkdir(output_folder)
#Move all the files into folder
os.renames(output_file,         output_folder + '/1 - ' + output_file)
os.renames(output_status,       output_folder + '/2 - ' + output_status)
os.renames(output_deep_status,  output_folder + '/3 - ' + output_deep_status)
shutil.copyfile(input_file,     output_folder + '/4 - ' + input_file)

print(output_result_array)