# -*- coding: utf-8 -*-
# Alle delprogrammer som skal brukes av hovedprogrammet

from random import randint
import numpy as np

from Print_status import print_status
from openpyxl import load_workbook



def importer_stemmesedler(input_file):
    wb = load_workbook(input_file)
    ws = wb.active

    # Get names of all candidates
    candidates = []
    more_candidates = True
    R = 2
    C = 1

    while more_candidates:
        cand = ws.cell(row=R, column=C).value
        if cand is not None:
            candidates.append(cand)
            R = R + 1
        else:
            more_candidates = False

    # Get all votes
    ballots = []
    one_ballot = []
    more_ballots = True
    R = 2
    C = 2

    while more_ballots:
        bal = ws.cell(row=R, column=C).value
        if bal is not None:
            one_ballot.append(bal)
            R = R + 1
        else:
            R = 2
            C = C + 1
            ballots.append(one_ballot)
            one_ballot = []

            if ws.cell(row=R, column=C).value is None:
                more_ballots = False

    #Sett alle blanke stemmer til -1
    for i in range(len(ballots)):
        for j in range(len(ballots[i])):
            if ballots[i][j] == 0:
                ballots[i][j] = -1

    kandidater = candidates
    stemmesedler = ballots

    return_tuplet = kandidater, stemmesedler
    return return_tuplet


def finn_sperregrense(antall_kandidater, antall_stemmesedler):
    return (len(antall_stemmesedler) / len(antall_kandidater)) + 0.01


def forste_opptelling(stemmesedler, kandidater, out_status, out_deep_status):
    first_count = []
    score = []

    oversikt = np.zeros(shape=np.shape(stemmesedler))
    score = np.zeros(shape=np.shape(kandidater))

    for i in range(len(stemmesedler)):
        for j in range(len(stemmesedler[i])):
            if stemmesedler[i][j] is 1:
                score[j] = score[j] + 1
                oversikt[i][j] = -2

    for i in range(len(oversikt)):
        for j in range(len(oversikt[i])):
            if oversikt[i][j] < -1:
                for k in range(len(oversikt)):
                    if oversikt[k][j] > -2:
                        oversikt[k][j] = -1

    tekst = '\n\nFørste opptelling.\nKandidater med score:'
    print_status(out_status, tekst)
    print_status(out_deep_status, tekst)

    for i in range(len(kandidater)):
        name_spacing = ' ' * (len(max(kandidater, key=len)) - len(kandidater[i]))
        tekst = str('\n' + kandidater[i]) + name_spacing + '\t:\t' + str(score[i])
        print_status(out_status, tekst)
        print_status(out_deep_status, tekst)

    return np.array(score), np.array(oversikt)


def alle_andre_opptellinger(stemmesedler, oversikt, score, sperregrense):


    for i in range(len(score)):
        if score[i] > sperregrense:
            for j in range(len(oversikt)):

                kand_stemme = len(stemmesedler[0]) + 1
                kand_stemme_pos = -1

                for k in range(len(oversikt[j])):
                    if oversikt[j][k] == 0:
                        if stemmesedler[j][k] < kand_stemme:
                            kand_stemme = stemmesedler[j][k]
                            kand_stemme_pos = k

                if i == kand_stemme_pos:
                    oversikt[j][kand_stemme_pos] = -2

    for i in range(len(oversikt)):
        for j in range(len(oversikt[i])):
            if oversikt[i][j] < -1:
                for k in range(len(oversikt)):
                    if oversikt[k][j] > -2:
                        oversikt[k][j] = -1

    return oversikt


def noen_har_vunnet(stemmesedler, sperregrense, score):
    for i in range(len(score)):
        if score[i] > sperregrense:
            return True
    return False


def kontroll_antall_kandidater_igjen(score):
    antall_kandidater_igjen = 0
    for i in range(len(score)):
        if score[i] >= 0:
            antall_kandidater_igjen += 1
    if antall_kandidater_igjen >= 1:
        return True
    else:
        return False


def flere_kandidater_har_vunnet(score, rangering, sperregrense):
    print('Kontrollerer for flere kandidater')
    for i in range(len(score)):
        if score[i] > sperregrense:
            if rangering[i] < 1:
                print('Fant ny kandidat!')
                return True
    print('Fant IKKE ny kandidat!')
    return False


def ny_fordeling_etter_vunnet(score, stemmesedler, sperregrense, rangering, kandidater, oversikt):
    # Finn ut hvor mange stemmer som skal fordeles per kandidat
    antall_stemmer_per_kand = np.zeros(shape=np.shape(kandidater))
    for i in range(len(oversikt)):
        antall_stemmer_per_kand[oversikt[i].argmin()] += 1

    # LEGG VINNERE I RESULTAT
    while flere_kandidater_har_vunnet(score, rangering, sperregrense):
        temp_tall = 0
        temp_pos = 0
        for i in range(len(score)):
            if score[i] > sperregrense:
                if rangering[i] < 1:
                    if score[i] > temp_tall:
                        temp_tall = score[i]
                        temp_pos = i

        legg_til_resultat = 1
        for i in range(len(rangering)):
            if rangering[i] >= legg_til_resultat:
                legg_til_resultat = rangering[i] + 1

        rangering[temp_pos] = legg_til_resultat

    # FORDEL OVERFLØDIGE STEMMER PER KANDIDAT
    for i in range(len(stemmesedler)):

        ny_kand_pos = -1
        ny_kand_rang = 1000

        for j in range(len(stemmesedler[i])):
            if stemmesedler[i][j] < ny_kand_rang:
                if oversikt[i][j] > -1:
                    ny_kand_rang = stemmesedler[i][j]
                    ny_kand_pos = j

        # Hvis ny_kand_pos er høyere enn -1 betyr det at det er en kandidat som skal få foredeling
        if oversikt[i][ny_kand_pos] > -1:
            if oversikt[i][oversikt[i].argmin()] < -1:
                score[ny_kand_pos] += (score[oversikt[i].argmin()] - sperregrense) / antall_stemmer_per_kand[oversikt[i].argmin()]

        oversikt[i][oversikt[i].argmin()] = -1

    # FJERN POENG TIL DE SOM HAR VUNNET FRA SCORE
    for i in range(len(rangering)):
        if rangering[i] > 0:
            score[i] = 0

    return score, oversikt, rangering


def ny_fordeling_etter_tap(score, stemmesedler, sperregrense, rangering, kandidater, oversikt):
    # Finn ut hvor mange stemmer som skal fordeles per kandidat
    antall_stemmer_per_kand = np.zeros(shape=np.shape(kandidater))
    for i in range(len(oversikt)):
        antall_stemmer_per_kand[oversikt[i].argmin()] += 1

    # LEGG VINNERE I RESULTAT
    while flere_kandidater_har_vunnet(score, rangering, sperregrense):
        temp_tall = 0
        temp_pos = 0
        for i in range(len(score)):
            if score[i] > sperregrense:
                if rangering[i] < 1:
                    if score[i] > temp_tall:
                        temp_tall = score[i]
                        temp_pos = i

        legg_til_resultat = 1
        for i in range(len(rangering)):
            if rangering[i] >= legg_til_resultat:
                legg_til_resultat = rangering[i] + 1

        rangering[temp_pos] = legg_til_resultat

    # FORDEL OVERFLØDIGE STEMMER PER KANDIDAT
    for i in range(len(stemmesedler)):

        ny_kand_pos = -1
        ny_kand_rang = 1000

        for j in range(len(stemmesedler[i])):
            if stemmesedler[i][j] < ny_kand_rang:
                if oversikt[i][j] > -1:
                    ny_kand_rang = stemmesedler[i][j]
                    ny_kand_pos = j

        # Hvis ny_kand_pos er høyere enn -1 betyr det at det er en kandidat som skal få foredeling
        if oversikt[i][ny_kand_pos] > -1:
            if oversikt[i][oversikt[i].argmin()] < -1:
                score[ny_kand_pos] += (score[oversikt[i].argmin()] - sperregrense) / antall_stemmer_per_kand[oversikt[i].argmin()]

        oversikt[i][oversikt[i].argmin()] = -1

    # FJERN POENG TIL DE SOM HAR VUNNET FRA SCORE
    for i in range(len(rangering)):
        if rangering[i] > 0:
            score[i] = 0

    return score, oversikt, rangering
